#!/usr/bin/env python3
"""
NMMS Tracking Report — Modern Desktop Client
Built with PySide6 (Qt6) for a professional, native look & feel.
"""

import sys
import os
import json
import webbrowser
import time
import platform
import subprocess
import traceback
from datetime import datetime
from io import BytesIO

# ── PySide6 (Qt6) ────────────────────────────────────────────────────────
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QStackedWidget, QLabel, QPushButton, QLineEdit,
    QTextEdit, QProgressBar, QFrame, QScrollArea, QDialog,
    QGraphicsDropShadowEffect
)
from PySide6.QtCore import (
    Qt, QThread, Signal, QTimer
)
from PySide6.QtGui import (
    QFont, QColor, QTextCursor, QTextCharFormat, QPainter,
    QLinearGradient, QBrush
)

# ── Business logic imports ───────────────────────────────────────────────
import requests
from bs4 import BeautifulSoup
from getmac import get_mac_address

# Excel
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image as PILImage

# Selenium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager


# ═══════════════════════════════════════════════════════════════════════════
# Utility helpers
# ═══════════════════════════════════════════════════════════════════════════

def copy_to_clipboard(text):
    """Copy text to clipboard."""
    try:
        import pyperclip
        pyperclip.copy(text)
        return True
    except ImportError:
        pass
    try:
        if platform.system() == 'Darwin':
            subprocess.run('pbcopy', input=text.encode('utf-8'), check=True)
        elif platform.system() == 'Windows':
            subprocess.run('clip', input=text.encode('utf-8'), check=True)
        else:
            subprocess.run(['xclip', '-selection', 'clipboard'], input=text.encode('utf-8'), check=True)
        return True
    except Exception:
        return False


def open_file_cross_platform(filepath):
    if platform.system() == 'Darwin':
        subprocess.call(('open', filepath))
    elif platform.system() == 'Windows':
        os.startfile(filepath)
    else:
        subprocess.call(('xdg-open', filepath))


# ═══════════════════════════════════════════════════════════════════════════
# Design System — colours, typography, spacing, and global QSS
# ═══════════════════════════════════════════════════════════════════════════

class Theme:
    # ── Palette ──────────────────────────────────────────────────────────
    PRIMARY         = "#0f172a"   # Deep navy
    PRIMARY_LIGHT   = "#1e293b"  # Slate
    ACCENT          = "#2563eb"   # Vibrant blue
    ACCENT_LIGHT    = "#3b82f6"   # Lighter blue
    ACCENT_DARK     = "#1d4ed8"   # Darker blue
    SUCCESS         = "#059669"   # Emerald
    SUCCESS_LIGHT   = "#10b981"   # Light emerald
    SUCCESS_DARK    = "#047857"   # Dark emerald
    WARNING         = "#d97706"   # Amber
    WARNING_LIGHT   = "#f59e0b"   # Light amber
    ERROR           = "#dc2626"   # Red
    ERROR_LIGHT     = "#ef4444"   # Light red
    ERROR_DARK      = "#b91c1c"   # Dark red
    SURFACE         = "#ffffff"   # Card bg
    BG              = "#f1f5f9"   # Page bg
    BG_ALT          = "#e2e8f0"   # Alt bg
    TEXT            = "#0f172a"   # Primary text
    TEXT_SEC        = "#64748b"   # Secondary text
    TEXT_MUTED      = "#94a3b8"   # Muted text
    BORDER          = "#e2e8f0"   # Borders
    BORDER_FOCUS    = "#93c5fd"   # Focus border
    LOG_BG          = "#0f172a"   # Log terminal bg
    LOG_TEXT        = "#e2e8f0"   # Log text
    SHADOW          = "rgba(15, 23, 42, 0.08)"  # Card shadow
    SHADOW_HOVER    = "rgba(15, 23, 42, 0.12)"  # Card shadow hover

    # ── Typography ───────────────────────────────────────────────────────
    FONT_FAMILY = "'SF Pro Display', 'Segoe UI', '.AppleSystemUIFont', system-ui, 'Inter', sans-serif"
    FONT_MONO   = "'JetBrains Mono', 'SF Mono', 'Fira Code', 'Consolas', monospace"

    # ── Spacing scale (px) ───────────────────────────────────────────────
    S2  = 2
    S4  = 4
    S6  = 6
    S8  = 8
    S10 = 10
    S12 = 12
    S16 = 16
    S20 = 20
    S24 = 24
    S32 = 32

    # ── Radii ────────────────────────────────────────────────────────────
    R4  = "4px"
    R6  = "6px"
    R8  = "8px"
    R10 = "10px"
    R12 = "12px"

    @classmethod
    def qss(cls):
        """Return the global QSS stylesheet."""
        return f"""
            QMainWindow, QWidget {{
                background-color: {cls.BG};
                font-family: {cls.FONT_FAMILY};
                font-size: 13px;
                color: {cls.TEXT};
            }}
            QScrollBar:vertical {{
                background-color: transparent;
                width: 8px;
                margin: 0;
            }}
            QScrollBar::handle:vertical {{
                background-color: {cls.TEXT_MUTED};
                border-radius: 4px;
                min-height: 30px;
            }}
            QScrollBar::handle:vertical:hover {{
                background-color: {cls.TEXT_SEC};
            }}
            QScrollBar::add-line:vertical,
            QScrollBar::sub-line:vertical {{
                height: 0;
            }}
            QScrollArea {{
                border: none;
                background: transparent;
            }}
        """


# ── Gradient-painted header widget ────────────────────────────────────────

class GradientHeader(QFrame):
    """A QFrame that paints a linear gradient background."""

    def __init__(self, color_start, color_end, parent=None):
        super().__init__(parent)
        self._start = QColor(color_start)
        self._end = QColor(color_end)
        self.setAttribute(Qt.WA_StyledBackground, False)

    def paintEvent(self, event):
        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing)
        g = QLinearGradient(0, 0, self.width(), self.height())
        g.setColorAt(0.0, self._start)
        g.setColorAt(1.0, self._end)
        p.fillRect(self.rect(), QBrush(g))
        p.end()


# # ── Styled widget factory functions ─────────────────────────────────────

def _btn(text, color=None, hover=None, **kwargs):
    """Return a modern QPushButton. If color is None, uses default blue."""
    c = color or Theme.ACCENT
    h = hover or Theme.ACCENT_LIGHT
    btn = QPushButton(text)
    btn.setCursor(Qt.PointingHandCursor)
    btn.setStyleSheet(f"""
        QPushButton {{
            background-color: {c};
            color: #ffffff;
            border: none;
            border-radius: {Theme.R8};
            padding: 10px 22px;
            font-size: 13px;
            font-weight: 600;
        }}
        QPushButton:hover {{
            background-color: {h};
        }}
        QPushButton:pressed {{
            background-color: {c};
        }}
        QPushButton:disabled {{
            background-color: {Theme.BORDER};
            color: {Theme.TEXT_MUTED};
        }}
    """)
    min_h = kwargs.pop("minimum_height", None)
    if min_h:
        btn.setMinimumHeight(min_h)
    return btn


def _label(text, size=13, weight=400, color=None):
    """Return a QLabel with the given typography."""
    lbl = QLabel(text)
    fnt = lbl.font()
    fnt.setPointSize(size)
    fnt.setWeight(QFont.Weight(weight))
    lbl.setFont(fnt)
    if color:
        lbl.setStyleSheet(f"color: {color}; background: transparent;")
    else:
        lbl.setStyleSheet("background: transparent;")
    return lbl


def _card(parent=None):
    """Return a QFrame styled as a white card with a drop-shadow."""
    card = QFrame(parent)
    card.setObjectName("modernCard")
    card.setStyleSheet(f"""
        QFrame#modernCard {{
            background-color: {Theme.SURFACE};
            border: 1px solid {Theme.BORDER};
            border-radius: {Theme.R12};
        }}
    """)
    # Drop shadow via graphics effect
    shadow = QGraphicsDropShadowEffect()
    shadow.setBlurRadius(24)
    shadow.setXOffset(0)
    shadow.setYOffset(2)
    shadow.setColor(QColor(Theme.SHADOW))
    card.setGraphicsEffect(shadow)
    return card


def _pill(text, color, bg=None):
    """Return a small inline QLabel styled as a pill/badge."""
    lbl = QLabel(text)
    b = bg or f"{color}20"
    lbl.setStyleSheet(f"""
        background-color: {b};
        color: {color};
        border-radius: {Theme.R4};
        padding: 2px 8px;
        font-size: 11px;
        font-weight: 600;
    """)
    return lbl


# ═══════════════════════════════════════════════════════════════════════════
# Server URL configuration
# ═══════════════════════════════════════════════════════════════════════════

DEFAULT_SERVER_URL = "https://nmms.palojori.in"
SERVER_URL = os.environ.get('NMMS_SERVER_URL')

if not SERVER_URL:
    config_paths = [
        os.path.join(os.path.dirname(os.path.abspath(__file__)), 'config.json'),
        os.path.expanduser('~/.nmms_config.json'),
    ]
    for cfg_path in config_paths:
        if os.path.exists(cfg_path):
            try:
                with open(cfg_path) as f:
                    config = json.load(f)
                    SERVER_URL = config.get('server_url')
                    if SERVER_URL:
                        break
            except Exception:
                pass

if not SERVER_URL:
    SERVER_URL = DEFAULT_SERVER_URL


# ═══════════════════════════════════════════════════════════════════════════
# Auth Manager (unchanged)
# ═══════════════════════════════════════════════════════════════════════════

class AuthManager:
    @staticmethod
    def get_mac():
        return get_mac_address()

    @staticmethod
    def check_license():
        mac = AuthManager.get_mac()
        try:
            res = requests.get(f"{SERVER_URL}/api/check_status?mac={mac}", timeout=5)
            return res.json()
        except Exception:
            return {"error": "Server unreachable. Check your internet or start the server."}


# ═══════════════════════════════════════════════════════════════════════════
# Excel Export Module (unchanged)
# ═══════════════════════════════════════════════════════════════════════════

class ExcelGenerator:
    @staticmethod
    def export(district, block, date, data, status_callback, stop_check):
        downloads = os.path.join(os.path.expanduser("~"), "Downloads")
        os.makedirs(downloads, exist_ok=True)
        filename = os.path.join(downloads, f"NMMS_Report_{block}_{date.replace('/', '-')}.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "NMMS Attendance"

        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Header
        ws.merge_cells('A1:M1')
        ws['A1'] = "NMMS TRACKING REPORT"
        ws['A1'].font = Font(size=18, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
        ws['A1'].alignment = center_align
        ws.row_dimensions[1].height = 35

        ws.merge_cells('A2:M2')
        ws['A2'] = f"Date: {date}  |  District: {district}  |  Block: {block}"
        ws['A2'].font = Font(size=12, bold=True, color="000000")
        ws['A2'].fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        ws['A2'].alignment = center_align
        ws.row_dimensions[2].height = 25

        ws.append([])

        headers = ["Sl. No.", "Panchayat", "Work Code", "MSR No.", "Work Name", "Worker Name", "Job Card", "Status", "Taken By", "Designation", "Geo Coordinates", "Photo 1", "Photo 2"]
        ws.append(headers)
        header_row_idx = 4

        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=header_row_idx, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 22
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 45
        ws.column_dimensions['F'].width = 25
        ws.column_dimensions['G'].width = 25
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 20
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 25
        ws.column_dimensions['L'].width = 22
        ws.column_dimensions['M'].width = 22

        row_idx = header_row_idx + 1
        sl_no = 1

        for item in data:
            if stop_check():
                break

            header_text = item.get('header_text', '')
            work_name = header_text.split('Work Name :')[-1].strip() if 'Work Name :' in header_text else header_text

            geo_coords = f"P1: {item.get('geo', 'N/A')}"
            if item.get('photo2_geo') and item.get('photo2_geo') != "N/A":
                geo_coords += f"\nP2: {item.get('photo2_geo')}"

            img1_obj, img2_obj = None, None
            if item.get('photo_url'):
                try:
                    res = requests.get(item['photo_url'])
                    img = PILImage.open(BytesIO(res.content))
                    img.thumbnail((140, 140))
                    img_path = f"temp_img1_{item['msr_no']}.png"
                    img.save(img_path)
                    img1_obj = img_path
                except Exception:
                    pass

            if item.get('photo2_url'):
                try:
                    res = requests.get(item['photo2_url'])
                    img = PILImage.open(BytesIO(res.content))
                    img.thumbnail((140, 140))
                    img_path = f"temp_img2_{item['msr_no']}.png"
                    img.save(img_path)
                    img2_obj = img_path
                except Exception:
                    pass

            panchayat = item.get('panchayat', '')
            for worker in item.get('workers', []):
                status_callback(f"Exporting row {sl_no} to Excel...", "info")
                ws.cell(row=row_idx, column=1, value=sl_no).alignment = center_align
                ws.cell(row=row_idx, column=2, value=panchayat).alignment = center_align
                ws.cell(row=row_idx, column=3, value=item.get('work_code', '')).alignment = center_align
                ws.cell(row=row_idx, column=4, value=item.get('msr_no', '')).alignment = center_align
                ws.cell(row=row_idx, column=5, value=work_name).alignment = center_align
                ws.cell(row=row_idx, column=6, value=worker['name']).alignment = center_align
                ws.cell(row=row_idx, column=7, value=worker['jobcard']).alignment = center_align

                status_cell = ws.cell(row=row_idx, column=8, value=worker['status'])
                status_cell.alignment = center_align
                status_cell.font = Font(bold=True, color="00B050" if "present" in worker['status'].lower() else "FF0000")

                ws.cell(row=row_idx, column=9, value=item.get('taken_by', 'N/A')).alignment = center_align
                ws.cell(row=row_idx, column=10, value=item.get('designation', 'N/A')).alignment = center_align
                ws.cell(row=row_idx, column=11, value=geo_coords).alignment = center_align
                ws.row_dimensions[row_idx].height = 100

                if img1_obj:
                    ws.add_image(OpenpyxlImage(img1_obj), f"L{row_idx}")
                if img2_obj:
                    ws.add_image(OpenpyxlImage(img2_obj), f"M{row_idx}")

                for col in range(1, 14):
                    ws.cell(row=row_idx, column=col).border = thin_border

                row_idx += 1
                sl_no += 1

        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=13)
        footer_cell = ws.cell(row=row_idx, column=1)
        footer_cell.value = "Report generated by Nrega Bot NMMS Tracker app"
        footer_cell.font = Font(italic=True, color="595959", bold=True, size=11)
        footer_cell.alignment = Alignment(horizontal="center", vertical="center")
        footer_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        ws.row_dimensions[row_idx].height = 25

        for col in range(1, 14):
            ws.cell(row=row_idx, column=col).border = thin_border

        wb.save(filename)

        for f in os.listdir():
            if f.startswith("temp_img") and f.endswith(".png"):
                try:
                    os.remove(f)
                except Exception:
                    pass

        return filename


# ═══════════════════════════════════════════════════════════════════════════
# Background scraper — QThread with signals (unchanged)
# ═══════════════════════════════════════════════════════════════════════════

class ScraperThread(QThread):
    log_signal        = Signal(str, str)   # message, level
    progress_signal   = Signal(float)      # 0.0–1.0
    progress_text_signal = Signal(str)
    finished_signal   = Signal(str)        # excel file path
    error_signal      = Signal(str, str)   # error message, traceback
    cancelled_signal  = Signal()

    def __init__(self, state, district, block, date):
        super().__init__()
        self.state   = state
        self.district = district
        self.block   = block
        self.date    = date
        self._stop   = False

    def stop(self):
        self._stop = True

    def run(self):
        driver = None
        try:
            self.log_signal.emit("Engine starting in headless mode...", "info")

            chrome_options = Options()
            chrome_options.add_argument("--headless=new")
            chrome_options.add_argument("--window-size=1920,1080")
            chrome_options.add_argument("--log-level=3")
            driver = webdriver.Chrome(
                service=Service(ChromeDriverManager().install()),
                options=chrome_options
            )
            wait = WebDriverWait(driver, 15)

            if self._stop:
                raise Exception("Process Cancelled by User.")

            self.log_signal.emit("Connecting to NREGA server...", "info")
            driver.get("https://mnregaweb4.dord.gov.in/netnrega/View_NMMS_atten_date_new.aspx?fin_year=2025-2026&&digest=HNrisV4bhHnb7Gve3mAKYQ")
            time.sleep(2)

            Select(wait.until(EC.presence_of_element_located((By.ID, "ctl00_ContentPlaceHolder1_ddlstate")))).select_by_visible_text(self.state)
            time.sleep(3)
            Select(wait.until(EC.presence_of_element_located((By.ID, "ctl00_ContentPlaceHolder1_ddl_attendance")))).select_by_visible_text(self.date)
            time.sleep(3)
            wait.until(EC.element_to_be_clickable((By.ID, "ctl00_ContentPlaceHolder1_btn_showreport"))).click()
            time.sleep(3)

            if self._stop:
                raise Exception("Process Cancelled by User.")

            self.log_signal.emit("Navigating to location...", "info")
            state_link = wait.until(EC.presence_of_element_located((
                By.XPATH,
                f"//a[translate(normalize-space(text()), 'abcdefghijklmnopqrstuvwxyz', 'ABCDEFGHIJKLMNOPQRSTUVWXYZ') = '{self.state}']"
            )))
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'}); arguments[0].click();", state_link)
            time.sleep(3)

            dist_link = wait.until(EC.presence_of_element_located((
                By.XPATH,
                f"//a[translate(normalize-space(text()), 'abcdefghijklmnopqrstuvwxyz', 'ABCDEFGHIJKLMNOPQRSTUVWXYZ') = '{self.district}']"
            )))
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'}); arguments[0].click();", dist_link)
            time.sleep(3)

            block_row = wait.until(EC.presence_of_element_located((
                By.XPATH,
                f"//tr[td[a[translate(normalize-space(text()), 'abcdefghijklmnopqrstuvwxyz', 'ABCDEFGHIJKLMNOPQRSTUVWXYZ') = '{self.block}'] or translate(normalize-space(text()), 'abcdefghijklmnopqrstuvwxyz', 'ABCDEFGHIJKLMNOPQRSTUVWXYZ') = '{self.block}']]"
            )))
            mr_count_link = block_row.find_element(By.XPATH, "./td[4]/a")
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'}); arguments[0].click();", mr_count_link)
            time.sleep(3)

            self.log_signal.emit("Compiling muster roll list...", "info")
            soup = BeautifulSoup(driver.page_source, 'html.parser')
            mr_list = []
            for row in soup.find('div', id='RepPr1').find('table').find_all('tr'):
                cols = row.find_all('td')
                if len(cols) >= 6 and cols[5].find('a'):
                    href = cols[5].find('a')['href']
                    mr_list.append({
                        "work_code": cols[4].text.strip(),
                        "msr_no": cols[5].find('a').text.strip(),
                        "mr_url": href if href.startswith('http') else "https://mnregaweb4.dord.gov.in/netnrega/" + href.lstrip('/')
                    })

            total_mrs = len(mr_list)
            self.log_signal.emit(f"Found {total_mrs} muster rolls to process.", "success")

            final_data = []
            for idx, mr in enumerate(mr_list):
                if self._stop:
                    raise Exception("Process Cancelled by User.")

                pct = (idx + 1) / total_mrs
                self.progress_signal.emit(pct)
                self.progress_text_signal.emit(f"Processing MR {idx + 1} of {total_mrs}...")
                self.log_signal.emit(f"Deep scraping MR {idx + 1}/{total_mrs} ({mr['msr_no']})...", "info")

                driver.get(mr['mr_url'])
                time.sleep(1.5)

                s = BeautifulSoup(driver.page_source, 'html.parser')

                def get_text(element_id):
                    el = s.find('span', id=element_id)
                    return el.text.strip() if el else "N/A"

                header_text_val = get_text('ctl00_ContentPlaceHolder1_lbl_dtl')
                panchayat_val = ''
                if 'Panchayat :' in header_text_val:
                    panchayat_val = header_text_val.split('Panchayat :')[-1].split('|')[0].strip()
                elif 'Gram Panchayat :' in header_text_val:
                    panchayat_val = header_text_val.split('Gram Panchayat :')[-1].split('|')[0].strip()

                mr.update({
                    "header_text": header_text_val,
                    "panchayat": panchayat_val,
                    "taken_by": get_text('ctl00_ContentPlaceHolder1_lbl_Taken_by'),
                    "designation": get_text('ctl00_ContentPlaceHolder1_lbl_Designation'),
                    "geo": get_text('ctl00_ContentPlaceHolder1_lbl_cordinates'),
                    "photo2_geo": get_text('ctl00_ContentPlaceHolder1_lbl_SecondCordinates'),
                    "workers": [
                        {"sno": r.find_all('td')[0].text.strip(),
                         "jobcard": r.find_all('td')[1].text.strip(),
                         "name": r.find_all('td')[2].text.strip(),
                         "status": r.find_all('td')[4].text.strip()}
                        for r in s.find('table', id='ctl00_ContentPlaceHolder1_Gridviewattandance').find_all('tr')[1:]
                    ] if s.find('table', id='ctl00_ContentPlaceHolder1_Gridviewattandance') else []
                })

                img1 = s.find('img', id='ctl00_ContentPlaceHolder1_img_groupPhoto')
                img2 = s.find('img', id='ctl00_ContentPlaceHolder1_img_SecondGroupPhoto')
                if img1:
                    src = img1.get('src')
                    mr['photo_url'] = src if src.startswith('http') else "https://mnregaweb4.dord.gov.in/netnrega/" + src.lstrip('/')
                if img2:
                    src = img2.get('src')
                    mr['photo2_url'] = src if src.startswith('http') else "https://mnregaweb4.dord.gov.in/netnrega/" + src.lstrip('/')

                final_data.append(mr)

            driver.quit()

            if not self._stop:
                self.progress_text_signal.emit("Generating Excel report...")
                self.log_signal.emit("Generating professional Excel report with photos...", "info")
                excel_file = ExcelGenerator.export(
                    self.district, self.block, self.date, final_data,
                    lambda msg, lvl="info": self.log_signal.emit(msg, lvl),
                    lambda: self._stop
                )

                if not self._stop:
                    self.progress_signal.emit(1.0)
                    self.progress_text_signal.emit("Complete!")
                    self.log_signal.emit("Success! Opening Excel file...", "success")
                    self.finished_signal.emit(excel_file)

        except Exception as e:
            if "Cancelled" in str(e):
                self.log_signal.emit("Scraping stopped by user.", "warning")
                self.progress_text_signal.emit("Cancelled")
                self.cancelled_signal.emit()
            else:
                tb_str = traceback.format_exc()
                self.log_signal.emit(f"Error: {str(e)}", "error")
                self.progress_text_signal.emit("Error occurred")
                self.error_signal.emit(str(e), tb_str)
            if driver:
                try:
                    driver.quit()
                except Exception:
                    pass


# ═══════════════════════════════════════════════════════════════════════════
# Main Window
# ═══════════════════════════════════════════════════════════════════════════

class NMMSApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("NMMS Tracking Report")
        self.setMinimumSize(680, 700)
        self.resize(780, 820)

        # State
        self.mac_id         = AuthManager.get_mac()
        self.user_info      = {}
        self.days_left      = 0
        self.stop_scraping  = False
        self._log_buffer    = []
        self._scraper       = None
        self._spinner_chars = ["⣾", "⣽", "⣻", "⢿", "⡿", "⣟", "⣯", "⣷"]

        # Apply global stylesheet
        self.setStyleSheet(Theme.qss())

        # Central stacked widget
        self.stack = QStackedWidget()
        self.setCentralWidget(self.stack)

        # Start with loading screen
        self._build_loading()
        QTimer.singleShot(500, self._do_license_check)

    # ── Screen builder helpers ─────────────────────────────────────────────

    def _clear_stack(self):
        while self.stack.count():
            w = self.stack.widget(0)
            self.stack.removeWidget(w)
            w.deleteLater()

    # ── Logging ────────────────────────────────────────────────────────────

    def log_message(self, message, level="info"):
        timestamp = datetime.now().strftime("%H:%M:%S")
        icons   = {"info":"\u2139", "success":"\u2713", "warning":"\u26a0", "error":"\u2717", "debug":"\u25b8"}
        colours = {"info": Theme.ACCENT, "success": Theme.SUCCESS, "warning": Theme.WARNING,
                   "error": Theme.ERROR, "debug": Theme.TEXT_MUTED}
        icon    = icons.get(level, "")
        colour  = colours.get(level, Theme.TEXT)
        formatted = f"[{timestamp}] {icon} {message}"

        self._log_buffer.append(formatted)

        if hasattr(self, 'log_text') and self.log_text is not None:
            cursor = self.log_text.textCursor()
            cursor.movePosition(QTextCursor.MoveOperation.End)
            fmt = QTextCharFormat()
            fmt.setForeground(QColor(colour))
            cursor.insertText(formatted + "\n", fmt)
            self.log_text.setTextCursor(cursor)
            self.log_text.ensureCursorVisible()

    def get_log_text(self):
        return "\n".join(self._log_buffer)

    def copy_log_to_clipboard(self):
        text = self.get_log_text() or "(No log entries)"
        ok = copy_to_clipboard(text)
        if ok:
            self.log_message("Log copied to clipboard", "success")

    # ── Loading screen ────────────────────────────────────────────────────

    def _build_loading(self):
        self._clear_stack()
        w = QWidget()
        lo = QVBoxLayout(w)
        lo.setAlignment(Qt.AlignCenter)
        lo.setSpacing(Theme.S12)

        # Spinning indicator
        self._spinner_label = _label("", size=36, color=Theme.ACCENT)
        self._spinner_label.setAlignment(Qt.AlignCenter)
        lo.addWidget(self._spinner_label)

        lo.addSpacing(Theme.S8)

        title = _label("NMMS Tracker", size=26, weight=700, color=Theme.PRIMARY)
        title.setAlignment(Qt.AlignCenter)
        lo.addWidget(title)

        sub = _label("Verifying License with Server", size=14, color=Theme.TEXT_SEC)
        sub.setAlignment(Qt.AlignCenter)
        lo.addWidget(sub)

        self._loading_error_container = QVBoxLayout()
        lo.addLayout(self._loading_error_container)

        self.stack.addWidget(w)
        self.stack.setCurrentWidget(w)

        self._spinner_idx = 0
        self._spinner_timer = QTimer(self)
        self._spinner_timer.timeout.connect(self._animate_spinner)
        self._spinner_timer.start(80)

    def _animate_spinner(self):
        if hasattr(self, '_spinner_label') and self._spinner_label is not None:
            ch = self._spinner_chars[self._spinner_idx % len(self._spinner_chars)]
            self._spinner_label.setText(ch)
            self._spinner_idx += 1

    def _stop_spinner(self):
        if hasattr(self, '_spinner_timer') and self._spinner_timer is not None:
            self._spinner_timer.stop()
            self._spinner_timer = None

    def _do_license_check(self):
        auth_data = AuthManager.check_license()
        self._stop_spinner()

        if "error" in auth_data:
            self._show_loading_error(auth_data['error'])
            return

        if not auth_data.get("registered"):
            self._build_registration()
        elif not auth_data.get("active"):
            self._build_expired()
        else:
            self.user_info = auth_data.get("user_data", {})
            self.days_left = auth_data.get("days_left", 0)
            self._build_dashboard()

    def _show_loading_error(self, msg):
        if hasattr(self, '_spinner_label') and self._spinner_label is not None:
            self._spinner_label.setText("\u2716")
            self._spinner_label.setStyleSheet(f"color: {Theme.ERROR}; background: transparent;")

        err_label = _label(msg, size=15, weight=600, color=Theme.ERROR)
        err_label.setAlignment(Qt.AlignCenter)
        err_label.setWordWrap(True)
        self._loading_error_container.addWidget(err_label)

        self._loading_error_container.addSpacing(Theme.S12)

        retry = _btn("\u21bb  Retry", Theme.ACCENT, Theme.ACCENT_LIGHT, minimum_height=44)
        retry.setMaximumWidth(200)
        retry_row = QHBoxLayout()
        retry_row.addStretch()
        retry_row.addWidget(retry)
        retry_row.addStretch()
        retry.clicked.connect(lambda: (self._clear_stack(), self._build_loading(),
                                        QTimer.singleShot(500, self._do_license_check)))
        self._loading_error_container.addLayout(retry_row)

    # ── Registration screen ───────────────────────────────────────────────

    def _build_registration(self):
        self._clear_stack()
        w = QWidget()
        lo = QVBoxLayout(w)
        lo.setAlignment(Qt.AlignCenter)

        card = _card()
        clo = QVBoxLayout(card)
        clo.setContentsMargins(Theme.S32, Theme.S32, Theme.S32, Theme.S32)
        clo.setSpacing(Theme.S8)
        clo.setAlignment(Qt.AlignCenter)

        clo.addWidget(_label("\U0001f44b", size=40))
        clo.addSpacing(Theme.S4)
        clo.addWidget(_label("Welcome to NMMS Tracker!", size=22, weight=700, color=Theme.PRIMARY))
        clo.addSpacing(Theme.S4)
        clo.addWidget(_label("Your device is not registered on the network.",
                             size=13, color=Theme.TEXT_SEC))
        clo.addWidget(_label("Register via the web portal to activate your 30-day trial.",
                             size=12, color=Theme.TEXT_MUTED))
        clo.addSpacing(Theme.S16)

        web_btn = _btn("\U0001f310  Open Registration Web Portal", Theme.ACCENT, Theme.ACCENT_LIGHT, minimum_height=48)
        web_btn.clicked.connect(lambda: webbrowser.open(f"{SERVER_URL}/register?mac={self.mac_id}"))
        clo.addWidget(web_btn)

        refresh_btn = _btn("\u21bb  I have Registered, Refresh Status", Theme.SUCCESS, Theme.SUCCESS_LIGHT, minimum_height=44)
        refresh_btn.clicked.connect(lambda: (self._clear_stack(), self._build_loading(),
                                              QTimer.singleShot(500, self._do_license_check)))
        clo.addWidget(refresh_btn)

        outer = QHBoxLayout()
        outer.addStretch()
        outer.addWidget(card)
        outer.addStretch()
        lo.addLayout(outer)

        self.stack.addWidget(w)
        self.stack.setCurrentWidget(w)

    # ── Expired screen ────────────────────────────────────────────────────

    def _build_expired(self):
        self._clear_stack()
        w = QWidget()
        lo = QVBoxLayout(w)
        lo.setAlignment(Qt.AlignCenter)

        card = _card()
        clo = QVBoxLayout(card)
        clo.setContentsMargins(Theme.S32, Theme.S32, Theme.S32, Theme.S32)
        clo.setSpacing(Theme.S8)
        clo.setAlignment(Qt.AlignCenter)

        clo.addWidget(_label("\u26a0", size=44))
        clo.addSpacing(Theme.S4)
        clo.addWidget(_label("Subscription Expired", size=22, weight=700, color=Theme.ERROR))
        clo.addSpacing(Theme.S4)
        clo.addWidget(_label(
            "Your 30-day trial or subscription has ended.\nPlease contact the administrator to renew.",
            size=13, color=Theme.TEXT_SEC
        ))
        clo.addSpacing(Theme.S16)

        retry_btn = _btn("\u21bb  Check Again", Theme.ERROR, Theme.ERROR_LIGHT, minimum_height=44)
        retry_btn.clicked.connect(lambda: (self._clear_stack(), self._build_loading(),
                                            QTimer.singleShot(500, self._do_license_check)))
        clo.addWidget(retry_btn)

        outer = QHBoxLayout()
        outer.addStretch()
        outer.addWidget(card)
        outer.addStretch()
        lo.addLayout(outer)

        self.stack.addWidget(w)
        self.stack.setCurrentWidget(w)

    # ── Main Dashboard ────────────────────────────────────────────────────

    def _build_dashboard(self):
        self._clear_stack()
        w = QWidget()
        outer_lo = QVBoxLayout(w)
        outer_lo.setContentsMargins(0, 0, 0, 0)
        outer_lo.setSpacing(0)

        # ── Header (gradient) ──────────────────────────────────────────────
        header = GradientHeader(Theme.PRIMARY, "#1a2744")
        header.setFixedHeight(78)
        hlo = QVBoxLayout(header)
        hlo.setContentsMargins(Theme.S24, Theme.S16, Theme.S24, Theme.S12)
        hlo.setSpacing(2)

        top_row = QHBoxLayout()
        top_row.addWidget(_label("NMMS Tracking Report", size=18, weight=700, color="#ffffff"))
        top_row.addStretch()

        # Status badge
        badge_color = Theme.SUCCESS
        badge = _pill(f"Connected \u2022 {self.days_left}d left", "#ffffff", "rgba(255,255,255,0.15)")
        top_row.addWidget(badge)
        hlo.addLayout(top_row)

        user_line = f"{self.user_info.get('name', 'User')}"
        hlo.addWidget(_label(user_line, size=11, weight=500, color="rgba(255,255,255,0.7)"))
        outer_lo.addWidget(header)

        # ── Location ribbon ─────────────────────────────────────────────────
        ribbon = GradientHeader(Theme.ACCENT, Theme.ACCENT_DARK)
        ribbon.setFixedHeight(46)
        rlo = QHBoxLayout(ribbon)
        rlo.setContentsMargins(Theme.S20, 0, Theme.S20, 0)

        loc = f"{self.user_info.get('state', '')}  |  {self.user_info.get('district', '')}  |  {self.user_info.get('block', '')}"
        self.loc_label = _label(f"\U0001f4cd  {loc}", size=13, weight=600, color="#ffffff")
        rlo.addWidget(self.loc_label)
        rlo.addStretch()

        edit_btn = QPushButton("\u270f  Edit")
        edit_btn.setCursor(Qt.PointingHandCursor)
        edit_btn.setFixedHeight(30)
        edit_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: rgba(255,255,255,0.15);
                color: #ffffff;
                border: 1px solid rgba(255,255,255,0.25);
                border-radius: 6px;
                padding: 4px 14px;
                font-size: 12px;
                font-weight: 600;
            }}
            QPushButton:hover {{
                background-color: rgba(255,255,255,0.25);
                border-color: rgba(255,255,255,0.4);
            }}
        """)
        edit_btn.clicked.connect(self._open_edit_popup)
        rlo.addWidget(edit_btn)
        outer_lo.addWidget(ribbon)

        # ── Scrollable body ─────────────────────────────────────────────────
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll.setStyleSheet("background: transparent;")
        body = QWidget()
        blo = QVBoxLayout(body)
        blo.setContentsMargins(Theme.S20, Theme.S16, Theme.S20, Theme.S8)
        blo.setSpacing(Theme.S12)

        # ── Row 1: Date + Actions side-by-side ────────────────────────────
        row1 = QHBoxLayout()
        row1.setSpacing(Theme.S12)

        # Card 1: Date Configuration
        card1 = _card()
        c1lo = QVBoxLayout(card1)
        c1lo.setContentsMargins(Theme.S16, Theme.S12, Theme.S16, Theme.S16)
        c1lo.setSpacing(Theme.S6)
        c1lo.addWidget(_label("\U0001f4c5  Date", size=14, weight=700))
        c1lo.addWidget(_label("Select the reporting date for extraction",
                              size=11, color=Theme.TEXT_MUTED))
        c1lo.addSpacing(Theme.S4)
        self.date_entry = QLineEdit()
        self.date_entry.setPlaceholderText("DD/MM/YYYY")
        self.date_entry.setText(datetime.now().strftime("%d/%m/%Y"))
        self.date_entry.setMinimumHeight(40)
        self.date_entry.setStyleSheet(f"""
            QLineEdit {{
                border: 1.5px solid {Theme.BORDER};
                border-radius: {Theme.R8};
                padding: 10px 14px;
                font-size: 14px;
                background: {Theme.SURFACE};
                color: {Theme.TEXT};
            }}
            QLineEdit:focus {{
                border-color: {Theme.ACCENT};
                background: #fafbff;
            }}
        """)
        c1lo.addWidget(self.date_entry)
        c1lo.addStretch()

        # Card 2: Actions
        card2 = _card()
        c2lo = QVBoxLayout(card2)
        c2lo.setContentsMargins(Theme.S16, Theme.S12, Theme.S16, Theme.S16)
        c2lo.setSpacing(Theme.S12)
        c2lo.addWidget(_label("\u25b6  Actions", size=14, weight=700))

        self.start_btn = _btn("  Start Extraction", Theme.ACCENT, Theme.ACCENT_LIGHT, minimum_height=46)
        self.start_btn.clicked.connect(self._start_scraping)
        c2lo.addWidget(self.start_btn)

        self.stop_btn = _btn("\u23f9  Stop", Theme.TEXT_MUTED, Theme.TEXT_MUTED, minimum_height=46)
        self.stop_btn.setEnabled(False)
        self.stop_btn.clicked.connect(self._stop_scraping)
        c2lo.addWidget(self.stop_btn)

        row1.addWidget(card1)
        row1.addWidget(card2)
        blo.addLayout(row1)

        # ── Card 3: Progress ────────────────────────────────────────────────
        card3 = _card()
        c3lo = QVBoxLayout(card3)
        c3lo.setContentsMargins(Theme.S16, Theme.S12, Theme.S16, Theme.S16)
        c3lo.setSpacing(Theme.S8)
        c3lo.addWidget(_label("\U0001f4ca  Progress", size=14, weight=700))

        # Progress bar container
        prog_container = QFrame()
        prog_container.setFixedHeight(32)
        prog_container.setStyleSheet("background: transparent;")
        prog_lo = QVBoxLayout(prog_container)
        prog_lo.setContentsMargins(0, 0, 0, 0)
        prog_lo.setSpacing(0)

        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        self.progress_bar.setTextVisible(False)
        self.progress_bar.setFixedHeight(8)
        self.progress_bar.setStyleSheet(f"""
            QProgressBar {{
                background-color: {Theme.BG_ALT};
                border-radius: 4px;
                border: none;
            }}
            QProgressBar::chunk {{
                background-color: {Theme.SUCCESS};
                border-radius: 4px;
            }}
        """)
        prog_lo.addWidget(self.progress_bar)
        c3lo.addWidget(prog_container)

        self.progress_label = _label("Ready", size=12, color=Theme.TEXT_SEC)
        c3lo.addWidget(self.progress_label)
        blo.addWidget(card3)

        # ── Card 4: Activity Log ────────────────────────────────────────────
        card4 = _card()
        c4lo = QVBoxLayout(card4)
        c4lo.setContentsMargins(Theme.S16, Theme.S12, Theme.S16, Theme.S16)
        c4lo.setSpacing(Theme.S8)

        log_header_row = QHBoxLayout()
        log_header_row.addWidget(_label("\U0001f4dd  Activity Log", size=14, weight=700))
        log_header_row.addStretch()
        copy_btn = QPushButton("\U0001f4cb  Copy Log")
        copy_btn.setCursor(Qt.PointingHandCursor)
        copy_btn.setFixedHeight(30)
        copy_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {Theme.PRIMARY_LIGHT};
                color: #ffffff;
                border: none;
                border-radius: {Theme.R6};
                padding: 4px 14px;
                font-size: 12px;
                font-weight: 600;
            }}
            QPushButton:hover {{
                background-color: {Theme.PRIMARY};
            }}
        """)
        copy_btn.clicked.connect(self.copy_log_to_clipboard)
        log_header_row.addWidget(copy_btn)
        c4lo.addLayout(log_header_row)

        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setMinimumHeight(160)
        self.log_text.setStyleSheet(f"""
            QTextEdit {{
                background-color: {Theme.LOG_BG};
                color: {Theme.LOG_TEXT};
                border: 1px solid {Theme.BORDER};
                border-radius: {Theme.R8};
                padding: 12px;
                font-family: {Theme.FONT_MONO};
                font-size: 12px;
                line-height: 1.5;
            }}
            QTextEdit:focus {{
                border-color: {Theme.ACCENT};
            }}
        """)
        c4lo.addWidget(self.log_text)
        blo.addWidget(card4)

        # Stretch
        blo.addStretch()

        scroll.setWidget(body)
        outer_lo.addWidget(scroll, stretch=1)

        # ── Footer ──────────────────────────────────────────────────────────
        footer = QFrame()
        footer.setStyleSheet(f"""
            QFrame {{
                background-color: {Theme.SURFACE};
                border: none;
                border-top: 1px solid {Theme.BORDER};
            }}
        """)
        footer.setFixedHeight(34)
        flo = QHBoxLayout(footer)
        flo.setContentsMargins(Theme.S20, 0, Theme.S20, 0)
        flo.addWidget(_label("Developed by Nrega Bot Team", size=10, color=Theme.TEXT_MUTED))
        flo.addStretch()
        flo.addWidget(_label("NMMS Tracker v2.0", size=10, color=Theme.TEXT_MUTED))
        outer_lo.addWidget(footer)

        self.stack.addWidget(w)
        self.stack.setCurrentWidget(w)

        self.log_message("System Ready. Waiting for instructions.", "info")

    # ── Edit popup ─────────────────────────────────────────────────────────

    def _open_edit_popup(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("Edit Configuration")
        dlg.setMinimumWidth(440)
        dlg.setModal(True)
        dlg.setStyleSheet(f"""
            QDialog {{
                background-color: {Theme.SURFACE};
            }}
        """)

        lo = QVBoxLayout(dlg)
        lo.setContentsMargins(0, 0, 0, 0)
        lo.setSpacing(0)

        # Gradient header
        hdr = GradientHeader(Theme.ACCENT, Theme.ACCENT_DARK)
        hdr.setFixedHeight(50)
        hlo = QHBoxLayout(hdr)
        hlo.setContentsMargins(Theme.S20, 0, Theme.S20, 0)
        hlo.addWidget(_label("\u270f  Location Configuration", size=15, weight=700, color="#ffffff"))
        hlo.addStretch()
        lo.addWidget(hdr)

        # Body
        body = QWidget()
        blo = QVBoxLayout(body)
        blo.setContentsMargins(Theme.S20, Theme.S16, Theme.S20, Theme.S20)
        blo.setSpacing(Theme.S10)

        blo.addWidget(_label("Update your location details for this session only.",
                             size=12, color=Theme.TEXT_SEC))

        def _make_entry(placeholder, default):
            e = QLineEdit()
            e.setPlaceholderText(placeholder)
            e.setText(default)
            e.setMinimumHeight(42)
            e.setStyleSheet(f"""
                QLineEdit {{
                    border: 1.5px solid {Theme.BORDER};
                    border-radius: {Theme.R8};
                    padding: 10px 14px;
                    font-size: 13px;
                    background: {Theme.SURFACE};
                    color: {Theme.TEXT};
                }}
                QLineEdit:focus {{
                    border-color: {Theme.ACCENT};
                    background: #fafbff;
                }}
            """)
            return e

        s_entry = _make_entry("State", self.user_info.get('state', ''))
        d_entry = _make_entry("District", self.user_info.get('district', ''))
        b_entry = _make_entry("Block", self.user_info.get('block', ''))
        blo.addWidget(s_entry)
        blo.addWidget(d_entry)
        blo.addWidget(b_entry)

        blo.addSpacing(Theme.S8)

        def save():
            self.user_info['state']    = s_entry.text().strip().upper()
            self.user_info['district'] = d_entry.text().strip().upper()
            self.user_info['block']    = b_entry.text().strip().upper()
            loc = f"{self.user_info['state']}  |  {self.user_info['district']}  |  {self.user_info['block']}"
            self.loc_label.setText(f"\U0001f4cd  {loc}")
            self.log_message("Location configuration updated.", "success")
            dlg.accept()

        btn_row = QHBoxLayout()
        btn_row.setSpacing(Theme.S10)
        cancel_btn = _btn("Cancel", Theme.TEXT_MUTED, Theme.TEXT_MUTED)
        cancel_btn.clicked.connect(dlg.reject)
        btn_row.addWidget(cancel_btn)

        save_btn = _btn("Save Details", Theme.SUCCESS, Theme.SUCCESS_LIGHT)
        save_btn.clicked.connect(save)
        btn_row.addWidget(save_btn)
        blo.addLayout(btn_row)

        lo.addWidget(body)
        dlg.exec()

    # ── Error dialog ──────────────────────────────────────────────────────

    def _show_error_dialog(self, title, message, traceback_str=None):
        dlg = QDialog(self)
        dlg.setWindowTitle("Error")
        dlg.setMinimumSize(560, 400)
        dlg.setModal(True)
        dlg.setStyleSheet(f"""
            QDialog {{
                background-color: {Theme.SURFACE};
            }}
        """)

        lo = QVBoxLayout(dlg)
        lo.setContentsMargins(0, 0, 0, 0)
        lo.setSpacing(0)

        # Red gradient header
        hdr = GradientHeader(Theme.ERROR, Theme.ERROR_DARK)
        hdr.setFixedHeight(50)
        hlo = QHBoxLayout(hdr)
        hlo.setContentsMargins(Theme.S20, 0, Theme.S20, 0)
        hlo.addWidget(_label(f"\u2717  {title}", size=15, weight=700, color="#ffffff"))
        hlo.addStretch()
        lo.addWidget(hdr)

        # Body
        body = QWidget()
        blo = QVBoxLayout(body)
        blo.setContentsMargins(Theme.S20, Theme.S16, Theme.S20, Theme.S20)
        blo.setSpacing(Theme.S10)

        blo.addWidget(_label("Error Details:", size=12, weight=600))

        error_area = QTextEdit()
        error_area.setReadOnly(True)
        error_area.setStyleSheet(f"""
            QTextEdit {{
                background-color: #fef2f2;
                border: 1px solid #fecaca;
                border-radius: {Theme.R8};
                padding: 12px;
                font-family: {Theme.FONT_MONO};
                font-size: 12px;
                color: {Theme.ERROR};
            }}
        """)
        error_area.setMinimumHeight(160)
        full = message
        if traceback_str:
            full += f"\n\n\u2501 Traceback \u2501\n{traceback_str}"
        error_area.setPlainText(full)
        blo.addWidget(error_area)

        btn_row = QHBoxLayout()
        btn_row.setSpacing(Theme.S10)

        copy_err = _btn("\U0001f4cb  Copy Error", Theme.PRIMARY_LIGHT, Theme.PRIMARY)
        copy_err.clicked.connect(lambda: (
            copy_to_clipboard(full),
            self.log_message("Error details copied to clipboard.", "success")
        ))
        btn_row.addWidget(copy_err)
        btn_row.addStretch()

        close_btn = _btn("Close", Theme.TEXT_MUTED, Theme.TEXT_MUTED)
        close_btn.clicked.connect(dlg.accept)
        btn_row.addWidget(close_btn)

        blo.addLayout(btn_row)
        lo.addWidget(body)

        dlg.exec()

    # ── Scraping lifecycle ────────────────────────────────────────────────

    def _start_scraping(self):
        s    = self.user_info.get('state')
        d    = self.user_info.get('district')
        b    = self.user_info.get('block')
        date = self.date_entry.text().strip()

        if not s or not d or not b or not date:
            self.log_message("Please set location details and date!", "error")
            return

        self.stop_scraping = False
        self.start_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
        self.stop_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {Theme.ERROR};
                color: #ffffff;
                border: none;
                border-radius: {Theme.R8};
                padding: 10px 22px;
                font-size: 13px;
                font-weight: 600;
            }}
            QPushButton:hover {{ background-color: {Theme.ERROR_LIGHT}; }}
            QPushButton:disabled {{ background-color: {Theme.BORDER}; color: {Theme.TEXT_MUTED}; }}
        """)

        self.progress_bar.setValue(0)
        self.progress_label.setText("Starting extraction...")
        self._log_buffer.clear()
        self.log_text.clear()

        self.log_message(f"Starting extraction for {s} > {d} > {b} on {date}", "info")

        self._scraper = ScraperThread(s, d, b, date)
        self._scraper.log_signal.connect(self.log_message)
        self._scraper.progress_signal.connect(lambda v: self.progress_bar.setValue(int(v * 100)))
        self._scraper.progress_text_signal.connect(self.progress_label.setText)
        self._scraper.finished_signal.connect(self._on_scrape_finished)
        self._scraper.error_signal.connect(self._on_scrape_error)
        self._scraper.cancelled_signal.connect(self._reset_buttons)
        self._scraper.start()

    def _stop_scraping(self):
        if self._scraper and self._scraper.isRunning():
            self._scraper.stop()
            self.log_message("Stopping process... Please wait.", "warning")
            self.stop_btn.setEnabled(False)

    def _on_scrape_finished(self, excel_file):
        self._reset_buttons()
        self.progress_bar.setValue(100)
        self.progress_label.setText("Complete!")
        self.log_message("Extraction complete. Opening report...", "success")

        def delayed_open():
            open_file_cross_platform(excel_file)
        QTimer.singleShot(500, delayed_open)

    def _on_scrape_error(self, err_msg, tb_str):
        self._reset_buttons()
        self.progress_label.setText("Error occurred")
        QTimer.singleShot(200, lambda: self._show_error_dialog(
            "Extraction Failed", err_msg, tb_str
        ))

    def _reset_buttons(self):
        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self.stop_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {Theme.TEXT_MUTED};
                color: #ffffff;
                border: none;
                border-radius: {Theme.R8};
                padding: 10px 22px;
                font-size: 13px;
                font-weight: 600;
            }}
            QPushButton:disabled {{ background-color: {Theme.BORDER}; color: {Theme.TEXT_MUTED}; }}
        """)

    # ── Legacy log_status (backward compat for ExcelGenerator callback) ────

    def log_status(self, text, color="#007bff"):
        level = "info"
        if color == "#28a745" or "success" in text.lower():
            level = "success"
        elif color == "#dc3545" or "error" in text.lower() or "\u2716" in text:
            level = "error"
        elif "warning" in text.lower() or "\u26a0" in text:
            level = "warning"
        self.log_message(text.replace("\u2705","").replace("\u274c","").replace("\u26a0\ufe0f","").replace("\u2139\ufe0f","").strip(), level)


# ═══════════════════════════════════════════════════════════════════════════
# Entry point
# ═══════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    app.setStyleSheet(Theme.qss())

    window = NMMSApp()
    window.show()
    sys.exit(app.exec())
