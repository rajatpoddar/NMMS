#!/usr/bin/env python3
"""
NMMS Tracking Report — Server-side Scraper Worker
Extracted from app.py, modified for Docker/web use with Remote WebDriver.
"""

import os
import json
import time
import uuid
import shutil
import threading
import traceback
from datetime import datetime
from io import BytesIO

# Selenium (Remote WebDriver for Docker)
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC

# Excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image as PILImage

# HTML parsing
from bs4 import BeautifulSoup

# HTTP for downloading photos
import requests


# ═══════════════════════════════════════════════════════════════════════════
# Configuration
# ═══════════════════════════════════════════════════════════════════════════

SELENIUM_URL = os.environ.get('SELENIUM_URL', 'http://selenium-chrome:4444/wd/hub')
OUTPUT_DIR = os.environ.get('NMMS_OUTPUT_DIR', '/app/outputs')
TASKS_DIR = os.environ.get('NMMS_TASKS_DIR', '/app/tasks')
MAX_CONCURRENT_TASKS = int(os.environ.get('MAX_CONCURRENT_TASKS', '5'))

NREGA_BASE_URL = "https://mnregaweb4.dord.gov.in/netnrega"
NREGA_NMMS_URL = (
    "https://mnregaweb4.dord.gov.in/netnrega/"
    "View_NMMS_atten_date_new.aspx?fin_year=2025-2026"
    "&&digest=HNrisV4bhHnb7Gve3mAKYQ"
)


# ═══════════════════════════════════════════════════════════════════════════
# Task Management (file-based, works across gunicorn workers)
# ═══════════════════════════════════════════════════════════════════════════

_tasks_lock = threading.Lock()
_active_tasks = {}  # task_id -> thread, for cancellation
_task_semaphore = threading.Semaphore(MAX_CONCURRENT_TASKS)  # limit concurrent scrapes


def init_dirs():
    """Ensure output and tasks directories exist and clean stale tasks."""
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    os.makedirs(TASKS_DIR, exist_ok=True)
    # Clean stale tasks from previous runs
    delete_old_tasks(max_age_hours=0)
    # Clean any orphaned temp directories from interrupted exports
    for item in os.listdir(OUTPUT_DIR):
        item_path = os.path.join(OUTPUT_DIR, item)
        if os.path.isdir(item_path) and item.startswith("temp_"):
            try:
                shutil.rmtree(item_path, ignore_errors=True)
            except Exception:
                pass
    # Process any queued tasks from a previous restart
    _process_queue()


def _task_file(task_id):
    return os.path.join(TASKS_DIR, f"{task_id}.json")


def read_task(task_id):
    """Read a task's status from its JSON file."""
    try:
        with open(_task_file(task_id)) as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return None


def write_task(task_data):
    """Write task status to JSON file. Thread-safe."""
    with _tasks_lock:
        task_id = task_data['task_id']
        with open(_task_file(task_id), 'w') as f:
            json.dump(task_data, f, indent=2, default=str)


def update_task_progress(task_id, pct, message, current=None, total=None):
    """Update progress in task file without overwriting other fields."""
    task = read_task(task_id)
    if not task:
        return
    task['state'] = 'running'
    task['progress']['pct'] = round(pct, 1)
    task['progress']['message'] = message
    if current is not None:
        task['progress']['current'] = current
    if total is not None:
        task['progress']['total'] = total
    write_task(task)


def append_task_log(task_id, message, level="info"):
    """Append a log entry to the task file."""
    task = read_task(task_id)
    if not task:
        return
    timestamp = datetime.now().strftime("%H:%M:%S")
    icons = {"info": "\u2139", "success": "\u2713", "warning": "\u26a0", "error": "\u2717"}
    icon = icons.get(level, "")
    task['log'].append(f"[{timestamp}] {icon} {message}")
    write_task(task)


def create_task(state, district, block, date):
    """Create a new task and return its ID."""
    task_id = str(uuid.uuid4())
    task = {
        'task_id': task_id,
        'state': 'pending',
        'params': {
            'state': state,
            'district': district,
            'block': block,
            'date': date,
        },
        'progress': {
            'pct': 0,
            'message': 'Queued...',
            'current': 0,
            'total': 0,
        },
        'log': [],
        'result': None,
        'error': None,
        'created_at': datetime.now().isoformat(),
        'completed_at': None,
    }
    write_task(task)
    return task_id


def delete_old_tasks(max_age_hours=24):
    """Clean up old task files and outputs."""
    now = time.time()
    for fname in os.listdir(TASKS_DIR):
        fpath = os.path.join(TASKS_DIR, fname)
        if os.path.isfile(fpath) and now - os.path.getmtime(fpath) > max_age_hours * 3600:
            try:
                task_id = fname.replace('.json', '')
                # Remove output file too
                for ext in ['.xlsx']:
                    opath = os.path.join(OUTPUT_DIR, f"{task_id}{ext}")
                    if os.path.exists(opath):
                        os.remove(opath)
                os.remove(fpath)
            except Exception:
                pass


# ═══════════════════════════════════════════════════════════════════════════
# Excel Generator (extracted from app.py, modified for server use)
# ═══════════════════════════════════════════════════════════════════════════

class ExcelGenerator:
    """Generates professional Excel reports with embedded photos."""

    @staticmethod
    def export(district, block, date, data, log_callback=None):
        """
        Generate Excel report and return path to saved file.
        log_callback: Optional function(message, level) for progress logging.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "NMMS Attendance"

        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # ── Title Row ──
        ws.merge_cells('A1:M1')
        ws['A1'] = "NMMS TRACKING REPORT"
        ws['A1'].font = Font(size=18, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
        ws['A1'].alignment = center_align
        ws.row_dimensions[1].height = 35

        # ── Subtitle ──
        ws.merge_cells('A2:M2')
        ws['A2'] = f"Date: {date}  |  District: {district}  |  Block: {block}"
        ws['A2'].font = Font(size=12, bold=True, color="000000")
        ws['A2'].fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        ws['A2'].alignment = center_align
        ws.row_dimensions[2].height = 25

        ws.append([])

        # ── Headers ──
        headers = [
            "Sl. No.", "Panchayat", "Work Code", "MSR No.", "Work Name",
            "Worker Name", "Job Card", "Status", "Taken By", "Designation",
            "Geo Coordinates", "Photo 1", "Photo 2"
        ]
        ws.append(headers)
        header_row_idx = 4

        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=header_row_idx, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        # ── Column widths ──
        widths = [8, 22, 25, 12, 45, 25, 25, 12, 20, 15, 25, 22, 22]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i) if i <= 26 else ''].width = w

        # ── Data rows ──
        row_idx = header_row_idx + 1
        sl_no = 1
        temp_dir = os.path.join(OUTPUT_DIR, f"temp_{uuid.uuid4().hex}")
        os.makedirs(temp_dir, exist_ok=True)

        try:
            for item_idx, item in enumerate(data):
                if log_callback:
                    log_callback(f"Exporting MR {item_idx + 1}/{len(data)}...", "info")

                header_text = item.get('header_text', '')
                work_name = (
                    header_text.split('Work Name :')[-1].strip()
                    if 'Work Name :' in header_text else header_text
                )

                geo_coords = f"P1: {item.get('geo', 'N/A')}"
                if item.get('photo2_geo') and item.get('photo2_geo') != "N/A":
                    geo_coords += f"\nP2: {item.get('photo2_geo')}"

                # Download photos
                img1_path, img2_path = None, None
                if item.get('photo_url'):
                    try:
                        r = requests.get(item['photo_url'], timeout=10)
                        img = PILImage.open(BytesIO(r.content))
                        img.thumbnail((140, 140))
                        p = os.path.join(temp_dir, f"img1_{item['msr_no']}.png")
                        img.save(p)
                        img1_path = p
                    except Exception:
                        pass

                if item.get('photo2_url'):
                    try:
                        r = requests.get(item['photo2_url'], timeout=10)
                        img = PILImage.open(BytesIO(r.content))
                        img.thumbnail((140, 140))
                        p = os.path.join(temp_dir, f"img2_{item['msr_no']}.png")
                        img.save(p)
                        img2_path = p
                    except Exception:
                        pass

                panchayat = item.get('panchayat', '')
                for worker in item.get('workers', []):
                    ws.cell(row=row_idx, column=1, value=sl_no).alignment = center_align
                    ws.cell(row=row_idx, column=2, value=panchayat).alignment = center_align
                    ws.cell(row=row_idx, column=3, value=item.get('work_code', '')).alignment = center_align
                    ws.cell(row=row_idx, column=4, value=item.get('msr_no', '')).alignment = center_align
                    ws.cell(row=row_idx, column=5, value=work_name).alignment = center_align
                    ws.cell(row=row_idx, column=6, value=worker['name']).alignment = center_align
                    ws.cell(row=row_idx, column=7, value=worker['jobcard']).alignment = center_align

                    status_cell = ws.cell(row=row_idx, column=8, value=worker['status'])
                    status_cell.alignment = center_align
                    is_present = "present" in worker['status'].lower()
                    status_cell.font = Font(bold=True, color="00B050" if is_present else "FF0000")

                    ws.cell(row=row_idx, column=9, value=item.get('taken_by', 'N/A')).alignment = center_align
                    ws.cell(row=row_idx, column=10, value=item.get('designation', 'N/A')).alignment = center_align
                    ws.cell(row=row_idx, column=11, value=geo_coords).alignment = center_align
                    ws.row_dimensions[row_idx].height = 100

                    if img1_path:
                        try:
                            ws.add_image(OpenpyxlImage(img1_path), f"L{row_idx}")
                        except Exception:
                            pass
                    if img2_path:
                        try:
                            ws.add_image(OpenpyxlImage(img2_path), f"M{row_idx}")
                        except Exception:
                            pass

                    for col in range(1, 14):
                        ws.cell(row=row_idx, column=col).border = thin_border

                    row_idx += 1
                    sl_no += 1

            # ── Footer ──
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=13)
            footer_cell = ws.cell(row=row_idx, column=1)
            footer_cell.value = "Report generated by Nrega Bot NMMS Tracker app"
            footer_cell.font = Font(italic=True, color="595959", bold=True, size=11)
            footer_cell.alignment = Alignment(horizontal="center", vertical="center")
            footer_cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            ws.row_dimensions[row_idx].height = 25

            for col in range(1, 14):
                ws.cell(row=row_idx, column=col).border = thin_border

            # ── Save ──
            filename = f"NMMS_Report_{block}_{date.replace('/', '-')}.xlsx"
            filepath = os.path.join(OUTPUT_DIR, filename)
            wb.save(filepath)
            return filepath

        finally:
            # Cleanup temp images
            if os.path.exists(temp_dir):
                shutil.rmtree(temp_dir, ignore_errors=True)


# ═══════════════════════════════════════════════════════════════════════════
# NREGA Scraper (extracted from app.py ScraperThread, using Remote WebDriver)
# ═══════════════════════════════════════════════════════════════════════════

class NregaScraper:
    """Scrapes NREGA NMMS attendance data using Selenium Remote WebDriver."""

    def __init__(self, state, district, block, date,
                 log_callback=None, progress_callback=None, stop_check=None):
        self.state = state.upper()
        self.district = district.upper()
        self.block = block.upper()
        self.date = date
        self.log = log_callback or (lambda msg, lvl="info": None)
        self.progress = progress_callback or (lambda pct, msg: None)
        self.stop_check = stop_check or (lambda: False)
        self.driver = None

    def _log(self, msg, level="info"):
        self.log(msg, level)

    def _progress(self, pct, msg):
        self.progress(pct, msg)

    def scrape(self):
        """Run the full scraping pipeline. Returns list of scraped muster roll data."""
        driver = None
        try:
            self._log("Engine starting in headless mode...", "info")

            chrome_options = Options()
            chrome_options.add_argument("--headless=new")
            chrome_options.add_argument("--window-size=1920,1080")
            chrome_options.add_argument("--log-level=3")
            chrome_options.add_argument("--no-sandbox")
            chrome_options.add_argument("--disable-dev-shm-usage")
            chrome_options.add_argument("--disable-gpu")

            self._log(f"Connecting to Selenium at {SELENIUM_URL}...", "info")
            driver = webdriver.Remote(
                command_executor=SELENIUM_URL,
                options=chrome_options
            )
            self.driver = driver
            wait = WebDriverWait(driver, 20)

            if self.stop_check():
                raise Exception("Process Cancelled by User.")

            self._log("Connecting to NREGA server...", "info")
            driver.get(NREGA_NMMS_URL)
            time.sleep(2)

            # ── Select State ──
            Select(wait.until(EC.presence_of_element_located(
                (By.ID, "ctl00_ContentPlaceHolder1_ddlstate")
            ))).select_by_visible_text(self.state)
            time.sleep(3)

            # ── Select Date ──
            Select(wait.until(EC.presence_of_element_located(
                (By.ID, "ctl00_ContentPlaceHolder1_ddl_attendance")
            ))).select_by_visible_text(self.date)
            time.sleep(3)

            # ── Click Show Report ──
            wait.until(EC.element_to_be_clickable(
                (By.ID, "ctl00_ContentPlaceHolder1_btn_showreport")
            )).click()
            time.sleep(3)

            if self.stop_check():
                raise Exception("Process Cancelled by User.")

            # ── Click State Link ──
            self._log("Navigating to location...", "info")
            state_link = wait.until(EC.presence_of_element_located((By.XPATH,
                f"//a[translate(normalize-space(text()), "
                f"'abcdefghijklmnopqrstuvwxyz', "
                f"'ABCDEFGHIJKLMNOPQRSTUVWXYZ') = '{self.state}']"
            )))
            driver.execute_script(
                "arguments[0].scrollIntoView({block: 'center'}); arguments[0].click();",
                state_link
            )
            time.sleep(3)

            # ── Click District Link ──
            dist_link = wait.until(EC.presence_of_element_located((By.XPATH,
                f"//a[translate(normalize-space(text()), "
                f"'abcdefghijklmnopqrstuvwxyz', "
                f"'ABCDEFGHIJKLMNOPQRSTUVWXYZ') = '{self.district}']"
            )))
            driver.execute_script(
                "arguments[0].scrollIntoView({block: 'center'}); arguments[0].click();",
                dist_link
            )
            time.sleep(3)

            # ── Click Block MR Count ──
            block_row = wait.until(EC.presence_of_element_located((By.XPATH,
                f"//tr[td[a[translate(normalize-space(text()), "
                f"'abcdefghijklmnopqrstuvwxyz', "
                f"'ABCDEFGHIJKLMNOPQRSTUVWXYZ') = '{self.block}'] "
                f"or translate(normalize-space(text()), "
                f"'abcdefghijklmnopqrstuvwxyz', "
                f"'ABCDEFGHIJKLMNOPQRSTUVWXYZ') = '{self.block}']]"
            )))
            mr_count_link = block_row.find_element(By.XPATH, "./td[4]/a")
            driver.execute_script(
                "arguments[0].scrollIntoView({block: 'center'}); arguments[0].click();",
                mr_count_link
            )
            time.sleep(3)

            # ── Parse Muster Roll List ──
            self._log("Compiling muster roll list...", "info")
            soup = BeautifulSoup(driver.page_source, 'html.parser')
            mr_list = []
            table_div = soup.find('div', id='RepPr1')
            if table_div and table_div.find('table'):
                for row in table_div.find('table').find_all('tr'):
                    cols = row.find_all('td')
                    if len(cols) >= 6 and cols[5].find('a'):
                        href = cols[5].find('a')['href']
                        mr_list.append({
                            "work_code": cols[4].text.strip(),
                            "msr_no": cols[5].find('a').text.strip(),
                            "mr_url": (
                                href if href.startswith('http')
                                else f"{NREGA_BASE_URL}/" + href.lstrip('/')
                            )
                        })

            total_mrs = len(mr_list)
            self._log(f"Found {total_mrs} muster rolls to process.", "success")

            # ── Deep Scrape Each MR ──
            final_data = []
            for idx, mr in enumerate(mr_list):
                if self.stop_check():
                    raise Exception("Process Cancelled by User.")

                pct = (idx + 1) / total_mrs * 100
                self._progress(pct, f"Processing MR {idx + 1} of {total_mrs}...")
                self._log(f"Deep scraping MR {idx + 1}/{total_mrs} ({mr['msr_no']})...", "info")

                driver.get(mr['mr_url'])
                time.sleep(1.5)

                s = BeautifulSoup(driver.page_source, 'html.parser')

                def get_text(element_id):
                    el = s.find('span', id=element_id)
                    return el.text.strip() if el else "N/A"

                header_text_val = get_text('ctl00_ContentPlaceHolder1_lbl_dtl')

                # Panchayat is in lbl_msg, not lbl_dtl
                msg_text_val = get_text('ctl00_ContentPlaceHolder1_lbl_msg')
                panchayat_val = ''
                if 'Panchayat :' in msg_text_val:
                    panchayat_val = msg_text_val.split('Panchayat :')[-1].strip()
                elif 'Panchayat:' in msg_text_val:
                    panchayat_val = msg_text_val.split('Panchayat:')[-1].strip()

                mr.update({
                    "header_text": header_text_val,
                    "panchayat": panchayat_val,
                    "taken_by": get_text('ctl00_ContentPlaceHolder1_lbl_Taken_by'),
                    "designation": get_text('ctl00_ContentPlaceHolder1_lbl_Designation'),
                    "geo": get_text('ctl00_ContentPlaceHolder1_lbl_cordinates'),
                    "photo2_geo": get_text('ctl00_ContentPlaceHolder1_lbl_SecondCordinates'),
                    "workers": [
                        {
                            "sno": r.find_all('td')[0].text.strip(),
                            "jobcard": r.find_all('td')[1].text.strip(),
                            "name": r.find_all('td')[2].text.strip(),
                            "status": r.find_all('td')[4].text.strip(),
                        }
                        for r in s.find('table',
                            id='ctl00_ContentPlaceHolder1_Gridviewattandance'
                        ).find_all('tr')[1:]
                    ] if s.find('table',
                        id='ctl00_ContentPlaceHolder1_Gridviewattandance'
                    ) else []
                })

                # Photo URLs
                img1 = s.find('img', id='ctl00_ContentPlaceHolder1_img_groupPhoto')
                img2 = s.find('img', id='ctl00_ContentPlaceHolder1_img_SecondGroupPhoto')
                if img1:
                    src = img1.get('src')
                    mr['photo_url'] = (
                        src if src.startswith('http')
                        else f"{NREGA_BASE_URL}/" + src.lstrip('/')
                    )
                if img2:
                    src = img2.get('src')
                    mr['photo2_url'] = (
                        src if src.startswith('http')
                        else f"{NREGA_BASE_URL}/" + src.lstrip('/')
                    )

                final_data.append(mr)

            return final_data

        finally:
            if driver:
                try:
                    driver.quit()
                except Exception:
                    pass


# ═══════════════════════════════════════════════════════════════════════════
# Background Task Runner
# ═══════════════════════════════════════════════════════════════════════════

def run_extraction_task(task_id, state, district, block, date):
    """
    Main background function that runs scraping + Excel export.
    Updates task status via JSON files.
    """
    try:
        update_task_progress(task_id, 0, "Starting...")
        append_task_log(task_id, f"Starting extraction for {state} > {district} > {block} on {date}", "info")

        stop_check = lambda: read_task(task_id).get('state') == 'cancelled' if read_task(task_id) else False

        def log_cb(msg, level="info"):
            append_task_log(task_id, msg, level)

        def progress_cb(pct, msg):
            update_task_progress(task_id, round(pct, 1), msg)

        # ── Scrape ──
        scraper = NregaScraper(
            state, district, block, date,
            log_callback=log_cb,
            progress_callback=progress_cb,
            stop_check=stop_check
        )
        data = scraper.scrape()

        # Check cancellation
        if stop_check():
            raise Exception("Process Cancelled by User.")

        # ── Generate Excel ──
        update_task_progress(task_id, 90, "Generating Excel report...")
        append_task_log(task_id, "Generating professional Excel report with photos...", "info")

        filepath = ExcelGenerator.export(
            district, block, date, data,
            log_callback=log_cb
        )

        if stop_check():
            raise Exception("Process Cancelled by User.")

        filename = os.path.basename(filepath)

        # ── Mark Complete ──
        task = read_task(task_id)
        task['state'] = 'done'
        task['progress']['pct'] = 100
        task['progress']['message'] = 'Complete!'
        task['result'] = {
            'filename': filename,
            'filepath': filepath,
        }
        task['completed_at'] = datetime.now().isoformat()
        write_task(task)

        append_task_log(task_id, "Extraction complete! Report ready for download.", "success")

    except Exception as e:
        tb = traceback.format_exc()
        task = read_task(task_id)
        if task:
            error_msg = str(e)
            if "Cancelled" in error_msg:
                task['state'] = 'cancelled'
                task['progress']['message'] = 'Cancelled'
                append_task_log(task_id, "Extraction cancelled by user.", "warning")
            else:
                task['state'] = 'error'
                task['progress']['message'] = 'Error occurred'
                task['error'] = f"{error_msg}\n\n{tb}"
                append_task_log(task_id, f"Error: {error_msg}", "error")
            write_task(task)


def _try_start_task(task_id):
    """
    Try to start a pending/queued task immediately.
    Acquires semaphore, updates state to 'running', starts thread.
    Returns True if started, False if no slot available or task not found.
    """
    # Try to acquire a semaphore slot
    acquired = _task_semaphore.acquire(blocking=False)
    if not acquired:
        return False

    task = read_task(task_id)
    if not task or task['state'] not in ('pending', 'queued'):
        # Task no longer exists or already started/completed — release slot
        _task_semaphore.release()
        return False

    # Mark as running
    task['state'] = 'running'
    task['progress']['message'] = 'Starting...'
    write_task(task)

    params = task['params']

    def _run_and_release(tid, s, d, b, dt):
        """Run task, then release slot and try next queued task."""
        try:
            run_extraction_task(tid, s, d, b, dt)
        finally:
            _task_semaphore.release()
            _process_queue()  # Start the next queued task

    thread = threading.Thread(
        target=_run_and_release,
        args=(task_id, params['state'], params['district'], params['block'], params['date']),
        daemon=True
    )
    _active_tasks[task_id] = thread
    thread.start()
    return True


def _process_queue():
    """Scan all queued tasks and start the next one if a slot is available."""
    try:
        # Collect all queued tasks
        queued = []
        with _tasks_lock:
            for fname in os.listdir(TASKS_DIR):
                if not fname.endswith('.json'):
                    continue
                task = read_task(fname.replace('.json', ''))
                if task and task['state'] == 'queued':
                    queued.append(task)

        if not queued:
            return

        # Sort oldest first (FIFO)
        queued.sort(key=lambda t: t['created_at'])

        # Try to start the first queued task
        _try_start_task(queued[0]['task_id'])

    except Exception:
        traceback.print_exc()


def start_extraction(state, district, block, date):
    """
    Start extraction in background thread.
    If max concurrent tasks reached, the task is queued automatically.
    Returns task_id.
    """
    task_id = create_task(state, district, block, date)

    # Try to start immediately
    started = _try_start_task(task_id)
    if not started:
        # No slot — queue the task
        task = read_task(task_id)
        task['state'] = 'queued'
        task['progress']['message'] = 'Queued — waiting for a free slot...'
        write_task(task)
        append_task_log(task_id, 'All slots busy. Task queued and will start automatically when a slot opens.', 'info')

    return task_id


def cancel_extraction(task_id):
    """Mark a task as cancelled. Works for queued and running tasks."""
    task = read_task(task_id)
    if task and task['state'] in ('pending', 'queued', 'running'):
        task['state'] = 'cancelled'
        task['progress']['message'] = 'Cancelled'
        write_task(task)
        return True
    return False
